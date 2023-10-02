#!/usr/bin/env python
# coding: utf-8

# In[1]:

def mergesheet():
    
    import numpy as np
    import os, re, csv, tkinter, shutil
    import pandas as pd
    from itertools import compress
    from InROI import InROI
    from tkinter import filedialog
    from openpyxl import load_workbook
    import xlsxwriter
    import glob
    
    stim_folder = input('Input folder name which have STIMULUS in left ROI:')
    control_folder = input('Input folder name which have CONTROL in left ROI:')
    flag = input('if doing stimulus merging, input 1, control merging input 0:')


    stimpath = os.getcwd()+'/'+ stim_folder
    conpath = os.getcwd()+'/'+ control_folder
    # 模糊搜索的文件名，例如 "*.txt" 或 "file_*.jpg"
    file_pattern = "*ROI*"
    # *ROI* files path saved in list↓
    stimfiles = glob.glob(stimpath + "/" + file_pattern)
    confiles = glob.glob(conpath+'/'+file_pattern)
    #print('stimulus files:\n',stimfiles)
    #print('control files:\n',confiles)


    merge_sheet_name_list = []
    dataframes_list = []
    for i in range(len(stimfiles)):
        print('data merging for %s'%stimfiles[i])
        merge_sheet_name_list = []
        dataframes_list = []
        # put sheet name into list
        stimsheet = pd.read_excel(stimfiles[i],sheet_name=None)
        stimsheetnames = list(stimsheet.keys())
        consheet = pd.read_excel(confiles[i],sheet_name=None)
        consheetnames = list(consheet.keys())
        merge_file_name = stimfiles[i].split('\\')[-1]
        print(merge_file_name)
        #print(stimsheetnames)
        for j in range(int(len(stimsheetnames)/2)):
            pairmerge = pd.DataFrame()

            if flag == '1':
                
                print('Doing stimulus group merging---')
                stim_sheet_value = pd.read_excel(stimfiles[i],sheet_name=stimsheetnames[j])
                con_sheet_value = pd.read_excel(confiles[i],sheet_name=consheetnames[int(len(stimsheetnames)/2+j)])
                pairmerge = pd.concat([stim_sheet_value,con_sheet_value],axis = 0)
                output_file_name = '/' + merge_file_name.split('.')[0]+'_Merging_stimulus.xlsx'
            elif flag == '0':
                print('Doing control group merging---')
                stim_sheet_value = pd.read_excel(stimfiles[i],sheet_name=stimsheetnames[int(len(stimsheetnames)/2+j)])
                con_sheet_value = pd.read_excel(confiles[i],sheet_name=consheetnames[j])
                pairmerge = pd.concat([stim_sheet_value,con_sheet_value],axis = 0)
                output_file_name = '/' + merge_file_name.split('.')[0]+'_Merging_control.xlsx'

            merge_sheet_name = stimsheetnames[j].split('_')[0]
            merge_sheet_name_list.append(str(merge_sheet_name))
            dataframes_list.append(pairmerge)
        #print(len(merge_sheet_name_list))
        #print(len(dataframes_list))
        with pd.ExcelWriter(os.getcwd() + output_file_name, engine = 'xlsxwriter') as writer:
            for k in range(len(dataframes_list)):
                dataframes_list[k].to_excel(writer ,sheet_name = merge_sheet_name_list[k] ,index = False ,header = True)

    print('MERGING DONE!!!')        

    writer.save()
    writer.close()


